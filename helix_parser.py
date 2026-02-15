#!/usr/bin/env python3
"""
Helix Native Preset Parser
Parses exported .hlx preset files and generates a detailed signal chain catalog.

Usage:
    python3 helix_parser.py /path/to/hlx/folder
    python3 helix_parser.py /path/to/single_preset.hlx
    python3 helix_parser.py /path/to/hlx/folder --xlsx output.xlsx
    python3 helix_parser.py /path/to/hlx/folder --csv output.csv
"""

import json
import os
import sys
import csv
import re
from pathlib import Path
from collections import OrderedDict

# ─── Helix Model ID → Human-Readable Name Lookup ───
# Categories: Amp, Cab, Dist, Delay, Mod, Reverb, Comp, EQ, Filter, Pitch,
#             Synth, Wah, Vol, FX Loop, Send/Return, Looper, Mic, IR, Utility

MODEL_DB = {
    # ══════════════ AMPS ══════════════
    "HD2_AmpUSSmallTweed": ("Amp", "US Small Tweed", "Fender Champ"),
    "HD2_AmpUSDeluxeNrml": ("Amp", "US Deluxe Nrml", "Fender Deluxe Reverb (Normal)"),
    "HD2_AmpUSDeluxeNrm": ("Amp", "US Deluxe Nrm", "Fender Deluxe Reverb (Normal)"),
    "HD2_AmpUSDeluxeVib": ("Amp", "US Deluxe Vib", "Fender Deluxe Reverb (Vibrato)"),
    "HD2_AmpUSDoubleNrm": ("Amp", "US Double Nrm", "Fender Twin Reverb (Normal)"),
    "HD2_AmpUSDoubleNrml": ("Amp", "US Double Nrm", "Fender Twin Reverb (Normal)"),
    "HD2_AmpUSDoubleVib": ("Amp", "US Double Vib", "Fender Twin Reverb (Vibrato)"),
    "HD2_AmpUSPrincess": ("Amp", "US Princess", "Fender Princeton Reverb"),
    "HD2_AmpSoupPro": ("Amp", "Soup Pro", "Supro S6420"),
    "HD2_AmpStoneAge185": ("Amp", "Stone Age 185", "Gibson EH-185"),
    "HD2_AmpTweedBluesNrm": ("Amp", "Tweed Blues Nrm", "Fender Bassman (Normal)"),
    "HD2_AmpTweedBluesBrt": ("Amp", "Tweed Blues Brt", "Fender Bassman (Bright)"),
    "HD2_AmpMailOrderTwin": ("Amp", "Mail Order Twin", "Silvertone 1484"),
    "HD2_AmpVoltageQueen": ("Amp", "Voltage Queen", "Victoria Electro King"),
    "HD2_AmpTucknGo": ("Amp", "Tuck & Go", "Ampeg Jet J-20"),
    # MESA/Boogie
    "HD2_AmpCaliIVR1": ("Amp", "Cali IV Rhythm 1", "MESA/Boogie Mk IV (R1)"),
    "HD2_AmpCaliIVRhythm1": ("Amp", "Cali IV Rhythm 1", "MESA/Boogie Mk IV (R1)"),
    "HD2_AmpCaliIVRhythm2": ("Amp", "Cali IV Rhythm 2", "MESA/Boogie Mk IV (R2)"),
    "HD2_AmpCaliIVLead": ("Amp", "Cali IV Lead", "MESA/Boogie Mk IV (Lead)"),
    "HD2_AmpCaliRectifire": ("Amp", "Cali Rectifire", "MESA/Boogie Dual Rectifier"),
    "HD2_AmpCaliTexasCh1": ("Amp", "Cali Texas Ch1", "MESA/Boogie Lonestar (Ch1)"),
    "HD2_AmpCaliTexasCh2": ("Amp", "Cali Texas Ch2", "MESA/Boogie Lonestar (Ch2)"),
    "HD2_AmpCaliBass": ("Amp", "Cali Bass", "MESA/Boogie Bass 400+"),
    "HD2_AmpCali400Ch1": ("Amp", "Cali 400 Ch1", "MESA/Boogie Bass 400 (Ch1)"),
    "HD2_AmpCali400Ch2": ("Amp", "Cali 400 Ch2", "MESA/Boogie Bass 400+ (Ch2)"),
    # Marshall
    "HD2_AmpBritPlexi": ("Amp", "Brit Plexi", "Marshall Super Lead 100"),
    "HD2_AmpBritPlexiNrm": ("Amp", "Brit Plexi Nrm", "Marshall Super Lead 100 (Normal)"),
    "HD2_AmpBritPlexiBrt": ("Amp", "Brit Plexi Brt", "Marshall Super Lead 100 (Bright)"),
    "HD2_AmpBritPlexiJump": ("Amp", "Brit Plexi Jump", "Marshall Super Lead 100 (Jumped)"),
    "HD2_AmpBrit2204": ("Amp", "Brit 2204", "Marshall JCM800 2204"),
    "HD2_AmpBritJ45Nrm": ("Amp", "Brit J-45 Nrm", "Marshall JTM-45 (Normal)"),
    "HD2_AmpBritJ45Brt": ("Amp", "Brit J-45 Brt", "Marshall JTM-45 (Bright)"),
    "HD2_AmpBritTremBrt": ("Amp", "Brit Trem Brt", "Marshall JTM-50 (Bright)"),
    "HD2_AmpBritTremJump": ("Amp", "Brit Trem Jump", "Marshall JTM-50 (Jumped)"),
    "HD2_AmpBritP75Nrm": ("Amp", "Brit P-75 Nrm", "Park 75 (Normal)"),
    "HD2_AmpBritP75Brt": ("Amp", "Brit P-75 Brt", "Park 75 (Bright)"),
    # Vox / Matchless
    "HD2_AmpEssexA15": ("Amp", "Essex A-15", "Vox AC-15"),
    "HD2_AmpEssexA30": ("Amp", "Essex A-30", "Vox AC-30 (Top Boost)"),
    "HD2_AmpA30FawnNrm": ("Amp", "A30 Fawn Nrm", "Vox AC-30 Fawn (Normal)"),
    "HD2_AmpA30FawnBrt": ("Amp", "A30 Fawn Brt", "Vox AC-30 Fawn (Bright)"),
    "HD2_AmpMatchstickCh1": ("Amp", "Matchstick Ch1", "Matchless DC-30 (Ch1)"),
    "HD2_AmpMatchstickCh2": ("Amp", "Matchstick Ch2", "Matchless DC-30 (Ch2)"),
    "HD2_AmpMatchstickJump": ("Amp", "Matchstick Jump", "Matchless DC-30 (Jumped)"),
    # High Gain
    "HD2_AmpGermanMahadeva": ("Amp", "German Mahadeva", "Bogner Shiva"),
    "HD2_AmpGermanUbersonic": ("Amp", "German Ubersonic", "Bogner Überschall"),
    "HD2_AmpSoloLeadOD": ("Amp", "Solo Lead OD", "Soldano SLO-100 (Overdrive)"),
    "HD2_AmpSoloLeadClean": ("Amp", "Solo Lead Clean", "Soldano SLO-100 (Clean)"),
    "HD2_AmpSoloLeadCrunch": ("Amp", "Solo Lead Crunch", "Soldano SLO-100 (Crunch)"),
    "HD2_AmpANGLMeteor": ("Amp", "ANGL Meteor", "ENGL Fireball 100"),
    "HD2_AmpAnglMeteor": ("Amp", "ANGL Meteor", "ENGL Fireball 100"),
    "HD2_AmpDasBenzinMega": ("Amp", "Das Benzin Mega", "Diezel Herbert (Mega Ch)"),
    "HD2_AmpPVPanama": ("Amp", "PV Panama", "Peavey 5150"),
    "HD2_AmpPVVitriolLead": ("Amp", "PV Vitriol Lead", "Peavey 6505+ (Lead)"),
    "HD2_AmpRevvGenPurple": ("Amp", "Revv Gen Purple", "Revv Generator 120 (Purple)"),
    "HD2_AmpRevvGenRed": ("Amp", "Revv Gen Red", "Revv Generator 120 (Red)"),
    "HD2_AmpPlacaterClean": ("Amp", "Placater Clean", "Friedman BE-100 (Clean)"),
    "HD2_AmpPlacaterDirty": ("Amp", "Placater Dirty", "Friedman BE-100 (BE/HBE)"),
    "HD2_AmpDerailedIngrid": ("Amp", "Derailed Ingrid", "Trainwreck Express"),
    "HD2_AmpBusyOneCh1": ("Amp", "Busy One Ch1", "Diezel VH4 (Ch1)"),
    "HD2_AmpBusyOneCh2": ("Amp", "Busy One Ch2", "Diezel VH4 (Ch2)"),
    # Orange / Others
    "HD2_AmpMandarin80": ("Amp", "Mandarin 80", "Orange OR80"),
    "HD2_AmpMandarinRocker": ("Amp", "Mandarin Rocker", "Orange Rockerverb"),
    "HD2_AmpWhoWatt100": ("Amp", "WhoWatt 100", "Hiwatt DR103"),
    "HD2_AmpCartographer": ("Amp", "Cartographer", "Ben Adrian Cartographer"),
    "HD2_AmpGrammaticoNrm": ("Amp", "Grammatico Nrm", "Grammatico LaGrange (Normal)"),
    "HD2_AmpGrammaticoJump": ("Amp", "Grammatico Jump", "Grammatico LaGrange (Jumped)"),
    "HD2_AmpGSG100": ("Amp", "GSG 100", "Grammatico GSG100"),
    "HD2_AmpInterstateZed": ("Amp", "Interstate Zed", "Dr. Z Route 66"),
    "HD2_AmpDividedDuo": ("Amp", "Divided Duo", "Divided by 13 9/15"),
    "HD2_AmpMoonJump": ("Amp", "Moo)))n Jump", "Moonlight Amplification"),
    "HD2_AmpArchetypeClean": ("Amp", "Archetype Clean", "Paul Reed Smith Archon (Clean)"),
    "HD2_AmpArchetypeLead": ("Amp", "Archetype Lead", "Paul Reed Smith Archon (Lead)"),
    "HD2_AmpJazzRivet120": ("Amp", "Jazz Rivet 120", "Roland JC-120"),
    "HD2_AmpCosmicGlue": ("Amp", "Cosmic Glue", "Supro S6616"),
    # Line 6 Originals
    "HD2_AmpLine6Litigator": ("Amp", "Litigator", "Line 6 Original (Blackface-style)"),
    "HD2_AmpLine6Badonk": ("Amp", "Badonk", "Line 6 Original (High Gain)"),
    "HD2_AmpLine6Elektrik": ("Amp", "Elektrik", "Line 6 Original"),
    "HD2_AmpLine6Epicenter": ("Amp", "Epicenter", "Line 6 Original (Bass)"),
    "HD2_AmpLine6Epic": ("Amp", "Line 6 Epic", "Line 6 Original"),
    "HD2_AmpLine6Fatality": ("Amp", "Line 6 Fatality", "Line 6 Original (Fatality)"),
    "HD2_AmpLine6Elmsley": ("Amp", "Elmsley", "Line 6 Original"),
    "HD2_AmpLine6Ventoux": ("Amp", "Ventoux", "Line 6 Original"),
    "HD2_AmpLine62204Mod": ("Amp", "Line 6 2204 Mod", "Line 6 Modified JCM800"),
    # Bass
    "HD2_AmpSVT4Pro": ("Amp", "SVT-4 Pro", "Ampeg SVT-4 Pro"),
    "HD2_AmpSVTNrml": ("Amp", "SVT Nrm", "Ampeg SVT (Normal)"),
    "HD2_AmpSVTBright": ("Amp", "SVT Bright", "Ampeg SVT (Bright)"),
    "HD2_AmpSVBeastNrm": ("Amp", "SV Beast Nrm", "Ampeg SVT (Beast Mode)"),
    "HD2_AmpSVBeastBrt": ("Amp", "SV Beast Brt", "Ampeg SVT (Beast Mode, Bright)"),
    "HD2_AmpGCougar800": ("Amp", "G Cougar 800", "Gallien-Krueger 800RB"),
    "HD2_AmpAgua51": ("Amp", "Agua 51", "Aguilar DB 751"),
    "HD2_AmpAguaSledge": ("Amp", "Agua Sledge", "Aguilar Tone Hammer"),
    "HD2_AmpDelSol300": ("Amp", "Del Sol 300", "Sunn Coliseum 300"),
    "HD2_AmpWoodyBlue": ("Amp", "Woody Blue", "Acoustic 360"),
    # Preamps
    "HD2_PreampUSDoubleNrm": ("Preamp", "US Double Nrm Pre", "Fender Twin Preamp"),
    "HD2_PreampUSDoubleVib": ("Preamp", "US Double Vib Pre", "Fender Twin Preamp (Vibrato)"),
    "HD2_PreampUSDeluxeNrm": ("Preamp", "US Deluxe Nrm Pre", "Fender Deluxe Reverb Preamp"),
    "HD2_PreampEssexA15": ("Preamp", "Essex A-15 Pre", "Vox AC-15 Preamp"),
    "HD2_PreampBusyOneCh2": ("Preamp", "Busy One Ch2 Pre", "Diezel VH4 Preamp (Ch2)"),
    "HD2_PreampSVT4Pro": ("Preamp", "SVT-4 Pro Pre", "Ampeg SVT-4 Pro Preamp"),
    "HD2_PreampVintagePre": ("Preamp", "Vintage Pre", "Generic Vintage Preamp"),
    "HD2_PreampCaliIVLead": ("Preamp", "Cali IV Lead Pre", "MESA/Boogie Mk IV Preamp"),
    "HD2_PreampBritPlexi": ("Preamp", "Brit Plexi Pre", "Marshall Plexi Preamp"),
    "HD2_PreampBrit2204": ("Preamp", "Brit 2204 Pre", "Marshall JCM800 Preamp"),

    # ══════════════ CABS (Standard) ══════════════
    "HD2_Cab1x10PrincessCopperhead": ("Cab", "1x10 Princess", "Fender Princeton"),
    "HD2_Cab1x12BlueBell": ("Cab", "1x12 Blue Bell", "Vox w/ Blue Alnico"),
    "HD2_Cab1x12Celest12H": ("Cab", "1x12 Celest 12-H", "Celestion G12H"),
    "HD2_Cab1x12DelSol": ("Cab", "1x12 Del Sol", "Sunn cab"),
    "HD2_Cab1x12FieldCoil": ("Cab", "1x12 Field Coil", "Field Coil speaker"),
    "HD2_Cab1x12Grammatico5E3": ("Cab", "1x12 Grammatico", "Grammatico 5E3"),
    "HD2_Cab1x12Lead80": ("Cab", "1x12 Lead 80", "Celestion Lead 80"),
    "HD2_Cab1x12MatchG25": ("Cab", "1x12 Match G25", "Matchless w/ Greenback 25"),
    "HD2_Cab1x12MatchH30": ("Cab", "1x12 Match H30", "Matchless w/ G12H30"),
    "HD2_Cab1x12PrincessBlue": ("Cab", "1x12 Princess Blue", "Fender Princeton w/ Blue"),
    "HD2_Cab1x12USDeluxe": ("Cab", "1x12 US Deluxe", "Fender Deluxe"),
    "HD2_Cab1x15TucknGo": ("Cab", "1x15 Tuck & Go", "Ampeg Jet cab"),
    "HD2_Cab1x18DelSol": ("Cab", "1x18 Del Sol", "Sunn 1x18"),
    "HD2_Cab1x18WoodyBlue": ("Cab", "1x18 Woody Blue", "Acoustic 360 cab"),
    "HD2_Cab2x12BlueBell": ("Cab", "2x12 Blue Bell", "Vox AC-30 w/ Blue Alnico"),
    "HD2_Cab2x12DoubleC12N": ("Cab", "2x12 Double C12N", "Fender Twin C12N"),
    "HD2_Cab2x12Interstate": ("Cab", "2x12 Interstate", "Dr. Z 2x12"),
    "HD2_Cab2x12JazzRivet": ("Cab", "2x12 Jazz Rivet", "Roland JC-120 cab"),
    "HD2_Cab2x12MailC12Q": ("Cab", "2x12 Mail C12Q", "Silvertone C12Q"),
    "HD2_Cab2x12SilverBell": ("Cab", "2x12 Silver Bell", "Vox w/ Silver Bell"),
    "HD2_Cab4x10Rhino": ("Cab", "4x10 Rhino", "Fender Super Reverb"),
    "HD2_Cab4x10TweedP10R": ("Cab", "4x10 Tweed P10R", "Fender Bassman 4x10"),
    "HD2_Cab4x121960T75": ("Cab", "4x12 1960 T75", "Marshall 1960 w/ T75"),
    "HD2_Cab4x12Blackback30": ("Cab", "4x12 Blackback 30", "Marshall w/ Blackback"),
    "HD2_Cab4x12Greenback25": ("Cab", "4x12 Greenback 25", "Marshall w/ Greenback 25W"),
    "HD2_Cab4X12CaliV30": ("Cab", "4x12 Cali V30", "MESA/Boogie w/ V30"),
    "HD2_Cab4x12CaliV30": ("Cab", "4x12 Cali V30", "MESA/Boogie w/ V30"),
    "HD2_Cab4x12UberT75": ("Cab", "4x12 Uber T75", "Bogner w/ T75"),
    "HD2_Cab4x12UberV30": ("Cab", "4x12 Uber V30", "Bogner w/ V30"),
    "HD2_Cab4x12WhoWatt100": ("Cab", "4x12 WhoWatt", "Hiwatt w/ Fane"),
    "HD2_Cab4x12XXLV30": ("Cab", "4x12 XXL V30", "Marshall 1960 w/ V30"),
    "HD2_Cab8x10SVBeast": ("Cab", "8x10 SV Beast", "Ampeg SVT 8x10"),
    "HD2_Cab1x6x9SoupProEllipse": ("Cab", "1x6x9 Soup Pro", "Supro S6420 Elliptical Speaker"),
    "HD2_Cab2x15Brute": ("Cab", "2x15 Brute", "Sunn 2x15 w/ JBL D140"),
    "HD2_Cab4x12Greenback20": ("Cab", "4x12 Greenback 20", "Marshall w/ Celestion G12M Greenback"),
    "HD2_Cab4x12MandarinEM": ("Cab", "4x12 Mandarin EM", "Orange PPC412"),
    "HD2_Cab4x12SoloLeadEM": ("Cab", "4x12 Solo Lead EM", "Soldano 4x12"),

    # ══════════════ CABS (Mic+IR "WithPan" variants) ══════════════
    "HD2_CabMicIr_1x10USPrincessWithPan": ("Cab", "1x10 US Princess", "Fender Princeton (dual mic)"),
    "HD2_CabMicIr_1x12CaliEXTWithPan": ("Cab", "1x12 Cali EXT", "MESA/Boogie Extension (dual mic)"),
    "HD2_CabMicIr_1x12Epicenter": ("Cab", "1x12 Epicenter", "Line 6 Epicenter cab"),
    "HD2_CabMicIr_1x12GrammaticoWithPan": ("Cab", "1x12 Grammatico", "Grammatico (dual mic)"),
    "HD2_CabMicIr_1x12OpenCast": ("Cab", "1x12 Open Cast", "Open-back 1x12"),
    "HD2_CabMicIr_1x12OpenCastWithPan": ("Cab", "1x12 Open Cast", "Open-back 1x12 (dual mic)"),
    "HD2_CabMicIr_1x12USDeluxe": ("Cab", "1x12 US Deluxe", "Fender Deluxe (mic)"),
    "HD2_CabMicIr_1x12USDeluxeWithPan": ("Cab", "1x12 US Deluxe", "Fender Deluxe (dual mic)"),
    "HD2_CabMicIr_2x12BlueBellWithPan": ("Cab", "2x12 Blue Bell", "Vox AC-30 Blue (dual mic)"),
    "HD2_CabMicIr_2x12DoubleC12N": ("Cab", "2x12 Double C12N", "Fender Twin (mic)"),
    "HD2_CabMicIr_2x12DoubleC12NWithPan": ("Cab", "2x12 Double C12N", "Fender Twin (dual mic)"),
    "HD2_CabMicIr_2x12JazzRivetWithPan": ("Cab", "2x12 Jazz Rivet", "Roland JC-120 (dual mic)"),
    "HD2_CabMicIr_2x12MailC12QWithPan": ("Cab", "2x12 Mail C12Q", "Silvertone (dual mic)"),
    "HD2_CabMicIr_2x12MandarinWithPan": ("Cab", "2x12 Mandarin", "Orange (dual mic)"),
    "HD2_CabMicIr_2x12MatchG25WithPan": ("Cab", "2x12 Match G25", "Matchless G25 (dual mic)"),
    "HD2_CabMicIr_2x12MatchH30WithPan": ("Cab", "2x12 Match H30", "Matchless H30 (dual mic)"),
    "HD2_CabMicIr_2x12SilverBellWithPan": ("Cab", "2x12 Silver Bell", "Vox Silver Bell (dual mic)"),
    "HD2_CabMicIr_2x15BruteWithPan": ("Cab", "2x15 Brute", "Bass 2x15 (dual mic)"),
    "HD2_CabMicIr_4x10GardenWithPan": ("Cab", "4x10 Garden", "Fender 4x10 (dual mic)"),
    "HD2_CabMicIr_4x10TweedP10RWithPan": ("Cab", "4x10 Tweed P10R", "Fender Bassman (dual mic)"),
    "HD2_CabMicIr_4x12BlackbackH30WithPan": ("Cab", "4x12 Blackback H30", "Marshall Blackback (dual mic)"),
    "HD2_CabMicIr_4x12CaliV30": ("Cab", "4x12 Cali V30", "MESA/Boogie V30 (mic)"),
    "HD2_CabMicIr_4x12CaliV30WithPan": ("Cab", "4x12 Cali V30", "MESA/Boogie V30 (dual mic)"),
    "HD2_CabMicIr_4x12Greenback25": ("Cab", "4x12 Greenback 25", "Marshall Greenback (mic)"),
    "HD2_CabMicIr_4x12Greenback25WithPan": ("Cab", "4x12 Greenback 25", "Marshall Greenback (dual mic)"),
    "HD2_CabMicIr_4x12MOONT75WithPan": ("Cab", "4x12 Moon T75", "Moon 4x12 T75 (dual mic)"),
    "HD2_CabMicIr_4x12Mandarin": ("Cab", "4x12 Mandarin", "Orange 4x12 (mic)"),
    "HD2_CabMicIr_4x12MandarinWithPan": ("Cab", "4x12 Mandarin", "Orange 4x12 (dual mic)"),
    "HD2_CabMicIr_4x12UberT75WithPan": ("Cab", "4x12 Uber T75", "Bogner T75 (dual mic)"),
    "HD2_CabMicIr_4x12UberV30WithPan": ("Cab", "4x12 Uber V30", "Bogner V30 (dual mic)"),
    "HD2_CabMicIr_8x10SVTAV": ("Cab", "8x10 SVT AV", "Ampeg SVT (mic)"),
    "HD2_CabMicIr_8x10SVTAVWithPan": ("Cab", "8x10 SVT AV", "Ampeg SVT (dual mic)"),

    # ══════════════ DISTORTION / DRIVE ══════════════
    "HD2_DistKinkyBoost": ("Drive", "Kinky Boost", "Xotic EP Booster"),
    "HD2_DistMinotaur": ("Drive", "Minotaur", "Klon Centaur"),
    "HD2_DistTeemah": ("Drive", "Teemah!", "Paul Cochrane Timmy"),
    "HD2_DistScream808": ("Drive", "Scream 808", "Ibanez TS808 Tube Screamer"),
    "HD2_DistHedgehogD9": ("Drive", "Hedgehog D9", "Maxon SD-9 Sonic Distortion"),
    "HD2_DistStuporOD": ("Drive", "Stupor OD", "Boss SD-1"),
    "HD2_DistDerangedMaster": ("Drive", "Deranged Master", "Dallas Rangemaster"),
    "HD2_DistTriangleFuzz": ("Drive", "Triangle Fuzz", "EHX Big Muff Pi (Triangle)"),
    "HD2_DistIndustrialFuzz": ("Drive", "Industrial Fuzz", "Z.Vex Fuzz Factory"),
    "HD2_DistBitcrusher": ("Drive", "Bitcrusher", "Line 6 Bitcrusher"),
    "HD2_DistHorizonDrive": ("Drive", "Horizon Drive", "Horizon Devices Precision Drive"),
    "HD2_DistDhyanaDrive": ("Drive", "Dhyana Drive", "Hermida Zendrive"),
    "HD2_DistCompulsiveDrive": ("Drive", "Compulsive Drive", "Fulltone OCD"),
    "HD2_DistValveDriver": ("Drive", "Valve Driver", "Chandler Tube Driver"),
    "HD2_DistTopSecretOD": ("Drive", "Top Secret OD", "DOD OD-250"),
    "HD2_DistDeezOneMod": ("Drive", "Deez One Mod", "Boss DS-1 (Keeley Mod)"),
    "HD2_DistMegaphone": ("Drive", "Megaphone", "Line 6 Megaphone"),
    "HD2_DistZeroAmpBassDI": ("Drive", "Zero Amp Bass DI", "Tech 21 SansAmp"),
    "HD2_DistHeirApparent": ("Drive", "Heir Apparent", "AnalogMan Prince of Tone"),
    "HD2_DistKWB": ("Drive", "KWB", "Benadrian KWB"),
    "HD2_DistToneSovereign": ("Drive", "Tone Sovereign", "Wampler Sovereign"),
    "HD2_DistTycoctaviaFuzz": ("Drive", "Tycoctavia Fuzz", "Tycobrahe Octavia"),
    "HD2_DistRamsHead": ("Drive", "Ram's Head", "EHX Big Muff Pi (Ram's Head)"),
    "HD2_DistObsidian7000": ("Drive", "Obsidian 7000", "Darkglass Microtubes B7K"),
    "HD2_DistClawthornDrive": ("Drive", "Clawthorn Drive", "EHX Crayon"),
    "HD2_DistArbitratorFuzz": ("Drive", "Arbitrator Fuzz", "Arbiter Fuzz Face"),
    "HD2_DistBallisticFuzz": ("Drive", "Ballistic Fuzz", "Balthazar Fuzz"),
    "HD2_DistPocketFuzz": ("Drive", "Pocket Fuzz", "Jordan Boss Tone"),
    "HD2_DistDeezOneVintage": ("Drive", "Deez One Vintage", "Boss DS-1"),
    "HD2_DistPillars": ("Drive", "Pillars", "Line 6 Pillars"),
    "HD2_DistVerminDist": ("Drive", "Vermin Dist", "Pro Co RAT"),
    "HD2_DistVitalDist": ("Drive", "Vital Dist", "MXR Dist+"),
    "HD2_DistThrifterFuzz": ("Drive", "Thrifter Fuzz", "Line 6 Thrifter"),
    "HD2_DistWringerFuzz": ("Drive", "Wringer Fuzz", "Garbage Wringer Fuzz"),
    "HD2_DistXenomorphFuzz": ("Drive", "Xenomorph Fuzz", "Industrial Fuzz variant"),
    "HD2_DistAmpegScramblerOD": ("Drive", "Ampeg Scrambler", "Ampeg Scrambler OD"),
    "HD2_DM4FacialFuzz": ("Drive", "Facial Fuzz", "Arbiter Fuzz Face (DM4)"),
    "HD2_DM4TubeDrive": ("Drive", "Tube Drive", "Chandler Tube Driver (DM4)"),

    # ══════════════ DELAY ══════════════
    "HD2_DelaySimpleDelay": ("Delay", "Simple Delay", "Line 6 Simple Delay"),
    "HD2_DelayHarmonyDelay": ("Delay", "Harmony Delay", "Line 6 Harmony Delay"),
    "HD2_DelayMultitap4": ("Delay", "Multitap 4", "Line 6 Multitap 4"),
    "HD2_DelayMultitap6": ("Delay", "Multitap 6", "Line 6 Multitap 6"),
    "HD2_DelayModChorusEcho": ("Delay", "Mod/Chorus Echo", "Line 6 Mod Delay"),
    "HD2_DelaySweepEcho": ("Delay", "Sweep Echo", "Line 6 Sweep Echo"),
    "HD2_DelayDualDelay": ("Delay", "Dual Delay", "Line 6 Dual Delay"),
    "HD2_DelayMultiPass": ("Delay", "Multipass", "Line 6 Multipass"),
    "HD2_DelayBucketBrigade": ("Delay", "Bucket Brigade", "Boss DM-2"),
    "HD2_DelayVintageDigitalV2": ("Delay", "Vintage Digital", "Roland RE-style digital"),
    "HD2_DelayTransistorTape": ("Delay", "Transistor Tape", "Maestro EP-3"),
    "HD2_DelayAdriaticDelay": ("Delay", "Adriatic Delay", "Boss DM-2w style"),
    "HD2_DelayElephantMan": ("Delay", "Elephant Man", "EHX Deluxe Memory Man"),
    "HD2_DelayPingPong": ("Delay", "Ping Pong", "Line 6 Ping Pong"),
    "HD2_DelayReverseDelay": ("Delay", "Reverse Delay", "Line 6 Reverse"),
    "HD2_DelayDuckedDelay": ("Delay", "Ducked Delay", "TC Electronic-style"),
    "HD2_DelayDoubleDouble": ("Delay", "Double Delay", "Line 6 Double Delay"),
    "HD2_DelayCosmosEcho": ("Delay", "Cosmos Echo", "Roland RE-201 Space Echo"),
    "HD2_DelayPitch": ("Delay", "Pitch Delay", "Line 6 Pitch Delay"),
    "HD2_DelaySwellAdriatic": ("Delay", "Swell Adriatic", "Auto-swell delay"),
    "HD2_DL4LowResDelay": ("Delay", "Low Res Delay", "Line 6 DL4 Lo Res"),
    "HD2_DL4PingPong": ("Delay", "DL4 Ping Pong", "Line 6 DL4 Ping Pong"),
    "Victoria_ShufflingDelay": ("Delay", "Shuffling Delay", "Line 6 Shuffling Delay"),

    # ══════════════ MODULATION ══════════════
    "HD2_TremoloTremolo": ("Mod", "Tremolo", "Line 6 Tremolo"),
    "HD2_TremoloPattern": ("Mod", "Pattern Tremolo", "Line 6 Pattern Tremolo"),
    "HD2_TremoloHarmonic": ("Mod", "Harmonic Tremolo", "Brownface-style"),
    "HD2_TremoloOpticalTrem": ("Mod", "Optical Trem", "Fender Optical Tremolo"),
    "HD2_Tremolo60sBiasTrem": ("Mod", "60s Bias Trem", "Vox Bias Tremolo"),
    "HD2_VibratoBubbleVibrato": ("Mod", "Bubble Vibrato", "Boss VB-2"),
    "HD2_Chorus": ("Mod", "Chorus", "Line 6 Chorus"),
    "HD2_Chorus70sChorus": ("Mod", "70s Chorus", "Boss CE-1"),
    "HD2_ChorusPlastiChorus": ("Mod", "PlastiChorus", "Arion SCH-Z"),
    "HD2_ChorusTrinityChorus": ("Mod", "Trinity Chorus", "Dytronics Tri-Stereo"),
    "HD2_FlangerGrayFlanger": ("Mod", "Gray Flanger", "MXR Flanger"),
    "HD2_FlangerHarmonicFlanger": ("Mod", "Harmonic Flanger", "A/DA Flanger"),
    "HD2_FlangerCourtesanFlange": ("Mod", "Courtesan Flange", "Electrix Flanger"),
    "HD2_PhaserScriptModPhase": ("Mod", "Script Mod Phase", "MXR Phase 90 (Script)"),
    "HD2_PhaserUbiquitousVibe": ("Mod", "Ubiquitous Vibe", "Shin-ei Uni-Vibe"),
    "HD2_RotaryRotary": ("Mod", "Rotary", "Leslie 145"),
    "HD2_Rotary145Rotary": ("Mod", "145 Rotary", "Leslie 145 Rotary"),
    "HD2_Rotary122Rotary": ("Mod", "122 Rotary", "Leslie 122 Rotary"),
    "HD2_RotaryVibeRotary": ("Mod", "Vibe Rotary", "Shin-ei Uni-Vibe"),
    "HD2_MM4AnalogFlanger": ("Mod", "Analog Flanger", "MXR Flanger (MM4)"),
    "HD2_MM4Dimension": ("Mod", "Dimension", "Roland Dimension D (MM4)"),
    "HD2_MM4UVibe": ("Mod", "U-Vibe", "Uni-Vibe (MM4)"),
    "HD2_RingModulatorAMRingMod": ("Mod", "AM Ring Mod", "Ring Modulator"),
    "HD2_RingModulatorPitchRingMod": ("Mod", "Pitch Ring Mod", "Pitch Ring Modulator"),

    # ══════════════ REVERB ══════════════
    "HD2_ReverbHall": ("Reverb", "Hall", "Line 6 Hall"),
    "HD2_ReverbPlate": ("Reverb", "Plate", "Line 6 Plate"),
    "HD2_ReverbRoom": ("Reverb", "Room", "Line 6 Room"),
    "HD2_ReverbTile": ("Reverb", "Tile", "Line 6 Tile"),
    "HD2_ReverbChamber": ("Reverb", "Chamber", "Line 6 Chamber"),
    "HD2_ReverbSpring": ("Reverb", "Spring", "Line 6 Spring"),
    "HD2_ReverbHxSpring": ("Reverb", "HX Spring", "Line 6 HX Spring"),
    "HD2_ReverbEcho": ("Reverb", "Echo", "Line 6 Echo"),
    "HD2_ReverbCave": ("Reverb", "Cave", "Line 6 Cave"),
    "HD2_ReverbDucking": ("Reverb", "Ducking", "Line 6 Ducking"),
    "HD2_ReverbOcto": ("Reverb", "Octo", "Line 6 Octo"),
    "HD2_ReverbGlitz": ("Reverb", "Glitz", "Line 6 Glitz"),
    "HD2_ReverbGanymede": ("Reverb", "Ganymede", "Line 6 Ganymede"),
    "HD2_ReverbSearchlights": ("Reverb", "Searchlights", "Line 6 Searchlights"),
    "HD2_ReverbPlateaux": ("Reverb", "Plateaux", "Line 6 Plateaux"),
    "HD2_ReverbDoubleTank": ("Reverb", "Double Tank", "Line 6 Double Tank"),
    "HD2_Reverb63Spring": ("Reverb", "'63 Spring", "Fender '63 Spring"),
    "HD2_ReverbParticle": ("Reverb", "Particle Verb", "Line 6 Particle Verb"),
    "VIC_ReverbDynAmbience": ("Reverb", "Dynamic Ambience", "Line 6 Dynamic Ambience"),
    "VIC_ReverbDynRoom": ("Reverb", "Dynamic Room", "Line 6 Dynamic Room"),
    "VIC_ReverbRotating": ("Reverb", "Rotating Reverb", "Line 6 Rotating Reverb"),
    "VIC_DynPlate": ("Reverb", "Dynamic Plate", "Line 6 Dynamic Plate"),

    # ══════════════ COMPRESSOR ══════════════
    "HD2_CompressorDeluxeComp": ("Comp", "Deluxe Comp", "Line 6 Deluxe Compressor"),
    "HD2_CompressorRedSqueeze": ("Comp", "Red Squeeze", "MXR Dyna Comp"),
    "HD2_CompressorKinkyComp": ("Comp", "Kinky Comp", "Xotic SP Compressor"),
    "HD2_CompressorLAStudioComp": ("Comp", "LA Studio Comp", "Teletronix LA-2A"),
    "HD2_Compressor3BandComp": ("Comp", "3-Band Comp", "Line 6 Multiband"),
    "HD2_CompressorRochesterComp": ("Comp", "Rochester Comp", "Ashly CLX-52"),
    "HD2_CompressorAutoSwell": ("Comp", "Auto Swell", "Line 6 Auto Swell"),

    # ══════════════ EQ ══════════════
    "HD2_EQParametric": ("EQ", "Parametric EQ", "Line 6 Parametric"),
    "HD2_EQGraphic10Band": ("EQ", "10-Band Graphic", "MXR 10-Band EQ"),
    "HD2_EQLowCutHighCut": ("EQ", "Low/High Cut", "Simple Filter"),
    "HD2_EQSimple3Band": ("EQ", "Simple EQ", "3-Band EQ"),
    "HD2_EQLowShelfHighShelf": ("EQ", "Low/High Shelf", "Shelf EQ"),
    "HD2_CaliQ": ("EQ", "Cali Q", "MESA/Boogie Graphic EQ"),

    # ══════════════ FILTER / WAH ══════════════
    "HD2_FilterAutoFilter": ("Filter", "Autofilter", "Line 6 Autofilter"),
    "HD2_FilterMutantFilter": ("Filter", "Mutant Filter", "Musitronics Mu-Tron III"),
    "HD2_FM4VoiceBox": ("Filter", "Voice Box", "Line 6 FM4 Voice Box"),
    "HD2_FilterMysterFilter": ("Filter", "Mystery Filter", "Mu-Tron III"),
    "HD2_FilterAshevillePattrn": ("Filter", "Asheville Pattrn", "Line 6 Asheville Pattern"),
    "HD2_WahTeardrop310": ("Wah", "Teardrop 310", "Dunlop Cry Baby"),
    "HD2_WahFassel": ("Wah", "Fassel", "Dunlop Cry Baby Original"),
    "HD2_WahChrome": ("Wah", "Chrome", "Vox V847"),
    "HD2_WahChromeCustom": ("Wah", "Chrome Custom", "Vox V847 Custom"),
    "HD2_WahWeeper": ("Wah", "Weeper", "Arbiter Cry Baby"),
    "HD2_WahConductor": ("Wah", "Conductor", "Maestro Boomerang"),
    "HD2_WahUKWah846": ("Wah", "UK Wah 846", "Vox V846"),
    "HD2_WahColorful": ("Wah", "Colorful", "Colorsound Wah"),
    "HD2_WahThroaty": ("Wah", "Throaty", "RMC Real McCoy"),

    # ══════════════ PITCH ══════════════
    "HD2_PitchSimplePitch": ("Pitch", "Simple Pitch", "Line 6 Pitch Shifter"),
    "HD2_PitchTwinHarmony": ("Pitch", "Twin Harmony", "Line 6 Harmonizer"),
    "HD2_PitchPitchWham": ("Pitch", "Pitch Wham", "Digitech Whammy"),
    "HD2_PitchDualPitch": ("Pitch", "Dual Pitch", "Line 6 Dual Pitch"),
    "VIC_PitchTwelveString": ("Pitch", "12-String", "Line 6 12-String Effect"),
    "HD2_DM4BassOctaver": ("Pitch", "Bass Octaver", "EBS OctaBass (DM4)"),

    # ══════════════ SYNTH ══════════════
    "HD2_Synth3NoteGenerator": ("Synth", "3 Note Generator", "Line 6 3-Note Synth"),
    "HD2_Synth4OSCGenerator": ("Synth", "4 OSC Generator", "Line 6 4-OSC Synth"),
    "HD2_SynthSubtractive": ("Synth", "Subtractive Synth", "Line 6 Subtractive Synth"),
    "HD2_FM4SynthOMatic": ("Synth", "Synth-O-Matic", "Line 6 FM4 Synth-O-Matic"),
    "HD2_FM4Growler": ("Synth", "Growler", "Line 6 FM4 Growler"),
    "HD2_FM4SynthString": ("Synth", "Synth String", "Line 6 FM4 Synth String"),
    "HD2_FM4TronDown": ("Synth", "Tron Down", "Mu-Tron III Down"),
    "HD2_FM4TronUp": ("Synth", "Tron Up", "Mu-Tron III Up"),

    # ══════════════ GATE ══════════════
    "HD2_GateNoiseGate": ("Gate", "Noise Gate", "Line 6 Noise Gate"),
    "HD2_GateHardGate": ("Gate", "Hard Gate", "Line 6 Hard Gate"),
    "HD2_GateHorizonGate": ("Gate", "Horizon Gate", "Horizon Devices Precision Gate"),

    # ══════════════ VOLUME / PAN / UTILITY ══════════════
    "HD2_VolPanVol": ("Utility", "Volume Pedal", "Volume Pedal"),
    "HD2_VolPanGain": ("Utility", "Gain Block", "Gain/Mute Block"),
    "HD2_VolPanPan": ("Utility", "Pan", "Pan Block"),
    "HD2_VolPanStereoImager": ("Utility", "Stereo Imager", "Stereo Width"),

    # ══════════════ LEGACY (L6SPB / Spearboard) ══════════════
    "L6SPB_12String": ("Pitch", "12-String", "Line 6 12-String Sim"),
    "L6SPB_AcousGtrSim": ("Utility", "Acoustic Sim", "Line 6 Acoustic Guitar Sim"),
    "L6SPB_InfSustain": ("Utility", "Infinite Sustain", "Line 6 Infinite Sustain"),
    "L6SPB_PolyChorus": ("Mod", "Poly Chorus", "Line 6 Poly Chorus"),
    "L6SPB_PolyDowntune": ("Pitch", "Poly Downtune", "Line 6 Poly Downtune"),
    "L6SPB_PolyPitch": ("Pitch", "Poly Pitch", "Line 6 Poly Pitch"),
    "L6SPB_PolyWham": ("Pitch", "Poly Wham", "Line 6 Poly Whammy"),

    # ══════════════ ROUTING / INFRASTRUCTURE ══════════════
    "HD2_AppDSPFlow1Input": ("Routing", "Input", "DSP Input"),
    "HD2_AppDSPFlow2Input": ("Routing", "Input B", "DSP Input B"),
    "HD2_AppDSPFlowOutput": ("Routing", "Output", "DSP Output"),
    "HD2_AppDSPFlowSplitY": ("Routing", "Split Y", "Y Split"),
    "HD2_AppDSPFlowSplitAB": ("Routing", "Split A/B", "A/B Split"),
    "HD2_AppDSPFlowSplitDynamic": ("Routing", "Split Dynamic", "Dynamic Split"),
    "HD2_AppDSPFlowSplitCrossover": ("Routing", "Split Crossover", "Crossover Split"),
    "HD2_AppDSPFlowJoin": ("Routing", "Join", "Path Join"),
}


def lookup_model(model_id):
    """Look up a model ID and return (category, name, based_on) or a parsed fallback."""
    if model_id in MODEL_DB:
        return MODEL_DB[model_id]
    # Try to parse the ID into something readable
    # e.g. HD2_AmpBritPlexi → "Amp: Brit Plexi"
    m = re.match(r'HD2_(\w+?)([A-Z][a-z].*)', model_id)
    if m:
        prefix = m.group(1)
        name = re.sub(r'([A-Z])', r' \1', m.group(2)).strip()
        cat_map = {
            'Amp': 'Amp', 'Preamp': 'Preamp', 'Cab': 'Cab', 'Dist': 'Drive',
            'Delay': 'Delay', 'Reverb': 'Reverb', 'Compressor': 'Comp',
            'EQ': 'EQ', 'Filter': 'Filter', 'Wah': 'Wah', 'Pitch': 'Pitch',
            'Synth': 'Synth', 'FM4': 'Synth', 'Tremolo': 'Mod', 'Chorus': 'Mod',
            'Flanger': 'Mod', 'Phaser': 'Mod', 'Rotary': 'Mod', 'VolPan': 'Utility',
            'Looper': 'Looper', 'FXLoop': 'FX Loop', 'App': 'Routing'
        }
        cat = cat_map.get(prefix, prefix)
        return (cat, f"{prefix} {name}", f"(Unknown: {model_id})")
    return ("Unknown", model_id, "")


def parse_hls_setlist(filepath):
    """Parse a .hls setlist file and return list of (preset_data, setlist_name, index) tuples."""
    import base64, zlib
    with open(filepath, 'r') as f:
        data = json.load(f)
    setlist_name = data.get('meta', {}).get('name', Path(filepath).stem)
    raw = base64.b64decode(data['encoded_data'])
    decompressed = zlib.decompress(raw)
    setlist = json.loads(decompressed.decode('utf-8'))
    presets_raw = setlist.get('presets', [])
    results = []
    for i, p in enumerate(presets_raw):
        wrapped = {'data': {'meta': p.get('meta', {}), 'tone': p.get('tone', {})}}
        results.append((wrapped, setlist_name, i))
    return results


def parse_preset(filepath, override_data=None, setlist_name=None, setlist_index=None):
    """Parse a single .hlx file (or pre-loaded data) and return structured data."""
    if override_data:
        data = override_data
    else:
        with open(filepath, 'r') as f:
            data = json.load(f)

    meta = data.get('data', {}).get('meta', {})
    tone = data.get('data', {}).get('tone', {})
    glob = tone.get('global', {})

    source = Path(filepath).name if not setlist_name else f"{setlist_name} #{setlist_index:03d}"
    preset_info = {
        'name': meta.get('name', Path(filepath).stem if not setlist_name else f'Preset {setlist_index}'),
        'file': source,
        'setlist': setlist_name or '',
        'setlist_index': setlist_index if setlist_index is not None else '',
        'tempo': glob.get('@tempo', ''),
        'topology0': glob.get('@topology0', ''),
        'topology1': glob.get('@topology1', ''),
        'snapshots': [],
        'dsp0': [],
        'dsp1': [],
    }

    # Extract snapshot names
    for i in range(8):
        snap = tone.get(f'snapshot{i}', {})
        if snap.get('@valid', False):
            preset_info['snapshots'].append(snap.get('@name', f'Snapshot {i}'))

    # Extract blocks from each DSP
    for dsp_name in ['dsp0', 'dsp1']:
        dsp = tone.get(dsp_name, {})
        blocks = []
        for key, val in dsp.items():
            if not isinstance(val, dict) or '@model' not in val:
                continue
            model_id = val['@model']
            # Skip routing infrastructure
            if model_id.startswith('HD2_AppDSP'):
                continue
            cat, name, based_on = lookup_model(model_id)
            blocks.append({
                'block': key,
                'position': val.get('@position', 99),
                'path': val.get('@path', 0),
                'enabled': val.get('@enabled', True),
                'model_id': model_id,
                'category': cat,
                'name': name,
                'based_on': based_on,
                'type': val.get('@type', ''),
                'stereo': val.get('@stereo', False),
            })
        blocks.sort(key=lambda b: (b['path'], b['position']))
        preset_info[dsp_name] = blocks

    return preset_info


def format_signal_chain(blocks):
    """Format blocks into a readable signal chain string."""
    if not blocks:
        return "(empty)"
    path0 = [b for b in blocks if b['path'] == 0]
    path1 = [b for b in blocks if b['path'] == 1]

    def chain_str(path_blocks):
        parts = []
        for b in path_blocks:
            status = "" if b['enabled'] else "[OFF] "
            parts.append(f"{status}{b['name']}")
        return " → ".join(parts)

    result = chain_str(path0)
    if path1:
        result += f"\n    Path B: {chain_str(path1)}"
    return result


def print_preset(info):
    """Pretty-print a parsed preset."""
    print(f"\n{'═' * 70}")
    print(f"  PRESET: {info['name']}")
    print(f"{'═' * 70}")
    if info['tempo']:
        print(f"  Tempo: {info['tempo']:.1f} BPM")
    if info['snapshots']:
        print(f"  Snapshots: {', '.join(info['snapshots'])}")

    # Identify key components
    all_blocks = info['dsp0'] + info['dsp1']
    amps = [b for b in all_blocks if b['category'] == 'Amp']
    cabs = [b for b in all_blocks if b['category'] == 'Cab']

    if amps:
        amp_strs = [a['name'] + ' (' + a['based_on'] + ')' for a in amps]
        print(f"  Amp(s): {', '.join(amp_strs)}")
    if cabs:
        cab_strs = [c['name'] + ' (' + c['based_on'] + ')' for c in cabs]
        print(f"  Cab(s): {', '.join(cab_strs)}")

    for dsp_name, label in [('dsp0', 'DSP 0'), ('dsp1', 'DSP 1')]:
        blocks = info[dsp_name]
        if blocks:
            print(f"\n  {label} Signal Chain:")
            chain = format_signal_chain(blocks)
            for line in chain.split('\n'):
                print(f"    {line}")

    # Summary table
    cats = {}
    for b in all_blocks:
        cats.setdefault(b['category'], []).append(b['name'])
    print(f"\n  Block Summary:")
    for cat in ['Amp', 'Preamp', 'Cab', 'Drive', 'Delay', 'Mod', 'Reverb',
                'Comp', 'EQ', 'Filter', 'Wah', 'Pitch', 'Synth', 'Utility',
                'FX Loop', 'Looper', 'Unknown']:
        if cat in cats:
            print(f"    {cat}: {', '.join(cats[cat])}")


def export_csv(presets, filepath):
    """Export presets to CSV."""
    with open(filepath, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow([
            'Preset Name', 'File', 'Tempo', 'Snapshots',
            'Amp(s)', 'Cab(s)', 'Drive(s)', 'Delay(s)', 'Mod(s)',
            'Reverb(s)', 'Comp(s)', 'Other Effects',
            'DSP0 Chain', 'DSP1 Chain',
            'Total Blocks', 'Categories Used'
        ])
        for info in presets:
            all_blocks = info['dsp0'] + info['dsp1']
            by_cat = {}
            for b in all_blocks:
                by_cat.setdefault(b['category'], []).append(b['name'])

            other = []
            for cat in ['EQ', 'Filter', 'Wah', 'Pitch', 'Synth', 'Utility',
                        'FX Loop', 'Looper', 'Unknown']:
                if cat in by_cat:
                    other.extend([f"[{cat}] {n}" for n in by_cat[cat]])

            writer.writerow([
                info['name'],
                info['file'],
                f"{info['tempo']:.1f}" if info['tempo'] else '',
                '; '.join(info['snapshots']),
                '; '.join(by_cat.get('Amp', [])),
                '; '.join(by_cat.get('Cab', [])),
                '; '.join(by_cat.get('Drive', [])),
                '; '.join(by_cat.get('Delay', [])),
                '; '.join(by_cat.get('Mod', [])),
                '; '.join(by_cat.get('Reverb', [])),
                '; '.join(by_cat.get('Comp', [])),
                '; '.join(other),
                format_signal_chain(info['dsp0']).replace('\n', ' | '),
                format_signal_chain(info['dsp1']).replace('\n', ' | '),
                len(all_blocks),
                ', '.join(sorted(by_cat.keys()))
            ])
    print(f"\nCSV exported to: {filepath}")


def export_xlsx(presets, filepath):
    """Export presets to a formatted Excel spreadsheet."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("openpyxl not installed. Install with: pip3 install openpyxl")
        print("Falling back to CSV export.")
        export_csv(presets, filepath.replace('.xlsx', '.csv'))
        return

    wb = Workbook()

    # ── Sheet 1: Overview ──
    ws = wb.active
    ws.title = "Preset Overview"

    header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    header_fill = PatternFill("solid", fgColor="333333")
    cat_fills = {
        'Amp': PatternFill("solid", fgColor="FFE0B2"),
        'Cab': PatternFill("solid", fgColor="FFF9C4"),
        'Drive': PatternFill("solid", fgColor="FFCDD2"),
        'Delay': PatternFill("solid", fgColor="B3E5FC"),
        'Mod': PatternFill("solid", fgColor="C8E6C9"),
        'Reverb': PatternFill("solid", fgColor="D1C4E9"),
        'Comp': PatternFill("solid", fgColor="F0F4C3"),
        'Synth': PatternFill("solid", fgColor="F8BBD0"),
    }

    headers = ['#', 'Preset Name', 'Setlist', 'Tempo', 'Amp(s)', 'Cab(s)',
               'Drive', 'Delay', 'Mod', 'Reverb', 'Comp',
               'Other', 'Total Blocks']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row, info in enumerate(presets, 2):
        all_blocks = info['dsp0'] + info['dsp1']
        by_cat = {}
        for b in all_blocks:
            by_cat.setdefault(b['category'], []).append(
                f"{b['name']} ({b['based_on']})" if b['based_on'] and 'Unknown' not in b['based_on'] else b['name']
            )

        other = []
        for cat in ['EQ', 'Filter', 'Wah', 'Pitch', 'Synth', 'Utility',
                    'FX Loop', 'Looper', 'Unknown']:
            if cat in by_cat:
                other.extend(by_cat[cat])

        ws.cell(row=row, column=1, value=row - 1)
        ws.cell(row=row, column=2, value=info['name']).font = Font(bold=True, name="Arial")
        ws.cell(row=row, column=3, value=info.get('setlist', ''))
        ws.cell(row=row, column=4, value=round(info['tempo'], 1) if info['tempo'] else None)
        ws.cell(row=row, column=5, value='\n'.join(by_cat.get('Amp', [])))
        ws.cell(row=row, column=6, value='\n'.join(by_cat.get('Cab', [])))
        ws.cell(row=row, column=7, value='\n'.join(by_cat.get('Drive', [])))
        ws.cell(row=row, column=8, value='\n'.join(by_cat.get('Delay', [])))
        ws.cell(row=row, column=9, value='\n'.join(by_cat.get('Mod', [])))
        ws.cell(row=row, column=10, value='\n'.join(by_cat.get('Reverb', [])))
        ws.cell(row=row, column=11, value='\n'.join(by_cat.get('Comp', [])))
        ws.cell(row=row, column=12, value='\n'.join(other))
        ws.cell(row=row, column=13, value=len(all_blocks))

        # Apply category color fills
        for col, cat in [(5, 'Amp'), (6, 'Cab'), (7, 'Drive'), (8, 'Delay'),
                         (9, 'Mod'), (10, 'Reverb'), (11, 'Comp'), (12, 'Synth')]:
            if cat in by_cat:
                ws.cell(row=row, column=col).fill = cat_fills.get(cat, PatternFill())

        # Wrap text
        for col in range(1, 14):
            ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical='top')

    # Column widths
    widths = [5, 24, 14, 8, 28, 24, 24, 20, 20, 18, 18, 28, 10]
    for i, w in enumerate(widths, 1):
        letter = chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)
        ws.column_dimensions[letter].width = w

    # ── Sheet 2: Signal Chains ──
    ws2 = wb.create_sheet("Signal Chains")
    headers2 = ['Preset', 'DSP', 'Position', 'Path', 'Block', 'Category',
                'Name', 'Based On', 'Enabled', 'Stereo']
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    row = 2
    for info in presets:
        for dsp_name in ['dsp0', 'dsp1']:
            for b in info[dsp_name]:
                ws2.cell(row=row, column=1, value=info['name'])
                ws2.cell(row=row, column=2, value=dsp_name.upper())
                ws2.cell(row=row, column=3, value=b['position'])
                ws2.cell(row=row, column=4, value=f"Path {'B' if b['path'] == 1 else 'A'}")
                ws2.cell(row=row, column=5, value=b['block'])
                ws2.cell(row=row, column=6, value=b['category'])
                ws2.cell(row=row, column=7, value=b['name'])
                ws2.cell(row=row, column=8, value=b['based_on'])
                ws2.cell(row=row, column=9, value='Yes' if b['enabled'] else 'No')
                ws2.cell(row=row, column=10, value='Stereo' if b['stereo'] else 'Mono')
                row += 1

    for i, w in enumerate([20, 8, 8, 8, 10, 10, 22, 28, 8, 8], 1):
        col_letter = chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)
        ws2.column_dimensions[col_letter].width = w

    wb.save(filepath)
    print(f"\nExcel spreadsheet exported to: {filepath}")


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 helix_parser.py <path> [--xlsx out.xlsx] [--csv out.csv]")
        print("  <path> can be a single .hlx file or a folder of .hlx files")
        sys.exit(1)

    target = sys.argv[1]
    xlsx_out = None
    csv_out = None

    # Parse optional args
    for i, arg in enumerate(sys.argv[2:], 2):
        if arg == '--xlsx' and i + 1 < len(sys.argv):
            xlsx_out = sys.argv[i + 1]
        elif arg == '--csv' and i + 1 < len(sys.argv):
            csv_out = sys.argv[i + 1]

    # Collect .hlx and .hls files
    hlx_files = []
    hls_files = []
    if os.path.isfile(target):
        if target.endswith('.hlx'):
            hlx_files = [target]
        elif target.endswith('.hls'):
            hls_files = [target]
        else:
            print(f"Error: {target} is not an .hlx or .hls file")
            sys.exit(1)
    elif os.path.isdir(target):
        hlx_files = sorted(Path(target).glob('*.hlx'))
        hls_files = sorted(Path(target).glob('*.hls'))
    else:
        print(f"Error: {target} not found")
        sys.exit(1)

    if not hlx_files and not hls_files:
        print(f"No .hlx or .hls files found in {target}")
        sys.exit(1)

    # Parse all presets
    presets = []

    # Parse .hls setlist files
    for fp in hls_files:
        try:
            entries = parse_hls_setlist(str(fp))
            print(f"Setlist: {Path(fp).stem} — {len(entries)} presets")
            for preset_data, sl_name, sl_idx in entries:
                info = parse_preset(str(fp), override_data=preset_data,
                                    setlist_name=sl_name, setlist_index=sl_idx)
                presets.append(info)
        except Exception as e:
            print(f"Error parsing setlist {fp}: {e}")

    # Parse individual .hlx files
    for fp in hlx_files:
        try:
            info = parse_preset(str(fp))
            presets.append(info)
        except Exception as e:
            print(f"Error parsing {fp}: {e}")

    print(f"\nTotal presets parsed: {len(presets)}\n")

    # Print to terminal
    for info in presets:
        print_preset(info)

    # Export if requested
    if xlsx_out:
        export_xlsx(presets, xlsx_out)
    if csv_out:
        export_csv(presets, csv_out)

    if not xlsx_out and not csv_out:
        print(f"\n{'─' * 70}")
        print(f"Tip: Add --xlsx catalog.xlsx or --csv catalog.csv to export")


if __name__ == '__main__':
    main()
